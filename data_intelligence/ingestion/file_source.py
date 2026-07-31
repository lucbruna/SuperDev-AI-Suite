"""File datasource ingestion (CSV, JSON, Excel)."""

from __future__ import annotations

import csv
import io
import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from data_intelligence.data_models import SourceType
from data_intelligence.ingestion.base import BaseSource


class FileSource(BaseSource):
    """Reads records from CSV, JSON or Excel files.

    * ``csv``  - first row is the header.
    * ``json`` - the file contains a list of objects, or an object with a
      ``data``/``results``/``items`` list.
    * ``xlsx`` - optional; requires ``openpyxl``.
    """

    source_type = SourceType.FILE

    def __init__(self, source_id: str, name: str, path: str | Path,
                 file_format: str = "auto",
                 base_dir: str | Path | None = None,
                 **config: Any) -> None:
        super().__init__(source_id, name, path=str(path),
                         file_format=file_format, **config)
        self.path = Path(path)
        self.base_dir = Path(base_dir) if base_dir else None
        self.file_format = (file_format if file_format != "auto"
                            else self.path.suffix.lstrip(".").lower() or "csv")

    def _contain(self, target: str | Path) -> Path:
        """Resolve *target* inside ``base_dir`` (CWE-22 path-traversal guard)."""
        candidate = Path(target)
        if self.base_dir is None:
            return candidate
        root = self.base_dir.resolve()
        resolved = candidate if candidate.is_absolute() else root / candidate
        resolved = resolved.resolve()
        if not resolved.is_relative_to(root):
            raise ValueError(
                f"path escapes base_dir for source '{self.name}': {target!r}")
        return resolved

    def fetch(self, source: Any = None) -> Iterable[dict[str, Any]]:  # noqa: ARG002
        self.path = self._contain(self.path)
        if not self.path.exists():
            raise FileNotFoundError(f"file not found: {self.path}")
        if self.file_format == "csv":
            return self._read_csv()
        if self.file_format in ("json", "jsonl"):
            return self._read_json()
        if self.file_format in ("xlsx", "xls"):
            return self._read_excel()
        raise ValueError(f"unsupported file format: {self.file_format}")

    def _read_csv(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        with self.path.open("r", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append(dict(row))
        return rows

    def _read_json(self) -> list[dict[str, Any]]:
        raw = json.loads(self.path.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            return [dict(item) for item in raw if isinstance(item, dict)]
        if isinstance(raw, dict):
            for key in ("data", "results", "items"):
                if key in raw and isinstance(raw[key], list):
                    return [dict(item) for item in raw[key]
                            if isinstance(item, dict)]
            return [raw]
        return []

    def _read_excel(self) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required to read Excel files") from exc
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        if workbook is None:
            raise RuntimeError("unable to open workbook")
        sheet = workbook.active
        if sheet is None:
            return []
        headers: list[str] = []
        rows: list[dict[str, Any]] = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                headers = [str(cell or "") for cell in row]
                continue
            rows.append({headers[j]: cell
                         for j, cell in enumerate(row)
                         if j < len(headers)})
        return rows

    def write_csv(self, records: Iterable[dict[str, Any]],
                  path: str | Path) -> dict[str, Any]:
        """Writes records to a CSV file (helper for sinks/testing)."""
        target = self._contain(path)
        rows = list(records)
        if not rows:
            target.write_text("", encoding="utf-8")
            return {"written": 0}
        keys = list(rows[0].keys())
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)
        target.write_text(buffer.getvalue(), encoding="utf-8")
        return {"written": len(rows)}
